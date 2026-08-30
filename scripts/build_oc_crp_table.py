"""Convert Reference/OC CRP Table.xlsx into data/oc_crp_table.json.

Sheet1 holds the human-readable tier policy, Sheet2 holds one block per crime:
    row 0: <tier> | <crime name> ... col H = success chance
    row 1: 'Role'   | role names (cols B-G)
    row 2: 'Weight' | role weights, col H = crime success chance
    row 3: 'CPR'    | minimum CPR per role

Usage:  python scripts/build_oc_crp_table.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Reference" / "OC CRP Table.xlsx"
TARGET = ROOT / "data" / "oc_crp_table.json"
SUPPLEMENTAL_CRIMES = ROOT / "data" / "oc_crp_table_gth.json"
USERSCRIPT = ROOT / "scripts" / "tampermonkey" / "Torn OC CRP Fit.user.js"
EMBED_PATTERN = re.compile(
    r"(/\* BEGIN EMBEDDED TABLE.*?\*/\n)(.*?)(\n\s*/\* END EMBEDDED TABLE \*/)",
    re.DOTALL,
)

ROLE_COLUMNS = range(1, 7)  # columns B-G
CHANCE_COLUMN = 7  # column H

# Legend from Sheet2 columns K-M: weight band -> minimum CPR expected of the holder.
WEIGHT_BANDS = [
    {"label": "Very Low", "min_weight": 0.0, "min_cpr": 40},
    {"label": "Low", "min_weight": 0.05, "min_cpr": 65},
    {"label": "Standard", "min_weight": 0.15, "min_cpr": 70},
    {"label": "High", "min_weight": 0.25, "min_cpr": 75},
]

# Tier-wide relaxations already baked into the per-role CPR values.
TIER_ADJUSTMENTS = {"8": -5, "9": -5, "10": -15}


def _clean(value: object) -> str:
    return str(value).replace("\xa0", " ").strip()


def parse_tier_policy(sheet) -> dict[str, dict[str, str]]:
    policy: dict[str, dict[str, str]] = {}
    current: str | None = None
    for tier, requirement in sheet.iter_rows(min_row=2, values_only=True):
        if requirement is None:
            continue
        text = _clean(requirement)
        if tier is not None:
            current = _clean(tier).lstrip("Tt")
            lines = text.split("\n")
            policy[current] = {"summary": lines[0], "notes": lines[1:]}
        elif current is not None:
            policy[current]["notes"].append(text)
    return policy


def parse_crimes(sheet) -> list[dict]:
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    crimes: list[dict] = []

    for index, row in enumerate(rows):
        if not isinstance(row[0], int):
            continue

        tier = row[0]
        name = _clean(row[1])
        role_row = rows[index + 1]
        weight_row = rows[index + 2]
        cpr_row = rows[index + 3]

        roles = []
        for col in ROLE_COLUMNS:
            if role_row[col] in (None, "Chance"):
                continue
            roles.append(
                {
                    "position": _clean(role_row[col]),
                    "weight": round(float(weight_row[col] or 0), 4),
                    "min_cpr": int(cpr_row[col]),
                }
            )

        crimes.append(
            {
                "name": name,
                "tier": tier,
                "success_chance": round(float(weight_row[CHANCE_COLUMN] or 0), 4),
                "roles": roles,
            }
        )

    return crimes


def merge_supplemental_crimes(crimes: list[dict]) -> list[dict]:
    """Keep structured low-tier crimes that are absent from the workbook."""
    supplemental = json.loads(SUPPLEMENTAL_CRIMES.read_text(encoding="utf-8"))
    existing_names = {str(crime["name"]).casefold() for crime in crimes}

    for crime in supplemental.get("crimes", []):
        if int(crime.get("tier", 0)) > 2 or str(crime.get("name", "")).casefold() in existing_names:
            continue

        copied = json.loads(json.dumps(crime))
        if int(copied["tier"]) == 2:
            for role in copied.get("roles", []):
                role["min_cpr"] = 60
        crimes.append(copied)

    return sorted(crimes, key=lambda crime: (int(crime["tier"]), str(crime["name"])))


def embed_in_userscript(payload: dict) -> None:
    """Keep the userscript's offline copy of the table in sync with the JSON."""
    compact = json.dumps(
        {k: v for k, v in payload.items() if k != "tier_policy"},
        separators=(",", ":"),
    )
    text = USERSCRIPT.read_text(encoding="utf-8")
    replacement = rf"\1  const EMBEDDED_TABLE = {compact.replace(chr(92), chr(92) * 2)};\3"
    updated, count = EMBED_PATTERN.subn(replacement, text)
    if not count:
        raise SystemExit("EMBEDDED TABLE markers not found in the userscript")
    USERSCRIPT.write_text(updated, encoding="utf-8")


def main() -> None:
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    payload = {
        "source": SOURCE.name,
        "weight_bands": WEIGHT_BANDS,
        "tier_adjustments": TIER_ADJUSTMENTS,
        "tier_policy": parse_tier_policy(workbook["Sheet1"]),
        "crimes": merge_supplemental_crimes(parse_crimes(workbook["Sheet2"])),
    }

    TARGET.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    embed_in_userscript(payload)
    print(f"Wrote {len(payload['crimes'])} crimes to {TARGET.relative_to(ROOT)}")
    print(f"Embedded table into {USERSCRIPT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
